import json
from msilib.schema import tables
from select import select
from tkinter import W
import pandas as pd
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import date
from datetime import timedelta
import re
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook



# ws = load_workbook('Prensa_historico/device_traffic_prensa.xlsx')

# sheet = ws['Sheet1']

# ws.save('Prensa_historico/device_traffic_prensa.xlsx')


jsonstr = '[{"hola":"dos", "tres":"4"}, {"hola":"dos", "tres":"4"}, {"hola":"dos", "tres":"4"}]' 

json_obj = json.loads(jsonstr)
